#!/usr/bin/env python3
"""
Sync libo-data.js, react-app/public/exercises.json, and react-app/public/workouts.json
from the source-of-truth Excel and the app's exercises.json.

Exercises: derived from libo-app-v2/src/data/exercises.json (already correct, 634 exercises).
Workouts: parsed from Exercise_Library_v5_Workouts.xlsx workout detail sheets.

Challenge Programs are excluded from workout output.
"""

import json
import re
import os
import openpyxl

REPO_ROOT = os.path.dirname(os.path.abspath(__file__))
MONO_ROOT = os.path.dirname(REPO_ROOT)

EXCEL_PATH = os.path.join(MONO_ROOT, "Brand-Management", "Exercises", "Exercise_Library_v5_Workouts.xlsx")
APP_EXERCISES_PATH = os.path.join(MONO_ROOT, "libo-app-v2", "src", "data", "exercises.json")

OUTPUT_LIBO_DATA = os.path.join(REPO_ROOT, "libo-data.js")
OUTPUT_EXERCISES_JSON = os.path.join(REPO_ROOT, "react-app", "public", "exercises.json")
OUTPUT_WORKOUTS_JSON = os.path.join(REPO_ROOT, "react-app", "public", "workouts.json")

# Workout category -> emoji mapping (matches existing libo-data.js)
WORKOUT_EMOJI = {
    "Gym": "💪",
    "Home": "🏠",
    "Cardio": "🏃",
    "Stretching": "🧘",
    "Morning Routine": "🌅",
}

# Workout detail sheets to parse (skip Challenge Programs)
WORKOUT_SHEETS = [
    "Gym Workouts",
    "Home Workouts",
    "Cardio Workouts",
    "Stretching Workouts",
    "Morning & Evening Routines",
]


def slugify(name: str) -> str:
    """Convert workout name to a slug ID like wk_chest_barbell_foundations."""
    s = name.lower()
    # Replace em dash and hyphens with space
    s = s.replace("—", " ").replace("–", " ").replace("-", " ")
    # Remove parentheses and their content for cleaner slugs
    s = re.sub(r"\(.*?\)", "", s)
    # Keep only alphanumeric and spaces
    s = re.sub(r"[^a-z0-9\s]", "", s)
    # Collapse whitespace and convert to underscores
    s = re.sub(r"\s+", "_", s.strip())
    # Remove trailing underscores
    s = s.rstrip("_")
    return "wk_" + s


def slugify_name(name: str) -> str:
    s = (name or "").lower()
    s = re.sub(r"[^a-z0-9]+", "_", s)
    return re.sub(r"_+", "_", s).strip("_")


def build_exercises():
    """Load exercises from the app's exercises.json and convert to libo-data.js format."""
    with open(APP_EXERCISES_PATH) as f:
        app_exercises = json.load(f)

    exercises = []
    for ex in app_exercises:
        out = {
            "id": ex["slug"],
            "name": ex["name"],
            "cat": ex["cat"],
            "bodyFocus": ex["bodyFocus"],
            "equipment": ex["equipment"],
            "machineRequired": ex["machineRequired"],
            "diff": ex["diff"],
            "variation": ex.get("variation", ""),
            "emoji": ex["emoji"],
            "setupNotes": ex.get("setupNotes", ""),
        }
        if ex.get("videoUrl"):
            out["videoUrl"] = ex["videoUrl"]
        if ex.get("videoUrlAlt"):
            out["videoUrlAlt"] = ex["videoUrlAlt"]
        # Carry parentId/parentName so the landing site can resolve variant → parent
        # (for thumbnail fallback — L/R variants share parent's thumb).
        # parentId is re-slugged to match landing's id scheme (slug-based, not gym_*/home_*).
        parent_name = ex.get("parentName") or ""
        if parent_name and ex["name"].startswith(parent_name):
            out["parentId"] = slugify_name(parent_name)
            out["parentName"] = parent_name
        exercises.append(out)

    return exercises


def parse_workout_sheet(ws):
    """Parse a workout detail sheet into a list of workout dicts."""
    workouts = []
    current = None

    for r in range(2, ws.max_row + 1):
        num = ws.cell(r, 1).value          # Column A: workout number
        name = ws.cell(r, 2).value          # Column B: workout name
        cat = ws.cell(r, 3).value           # Column C: category
        subcat = ws.cell(r, 4).value        # Column D: subcategory
        dur = ws.cell(r, 5).value           # Column E: duration
        diff = ws.cell(r, 6).value          # Column F: difficulty
        phase = ws.cell(r, 7).value         # Column G: phase
        exercise = ws.cell(r, 8).value      # Column H: exercise name
        sets = ws.cell(r, 9).value          # Column I: sets
        reps = ws.cell(r, 10).value         # Column J: reps/time

        # New workout header row (has a number and name)
        if num is not None and name is not None:
            if current is not None:
                workouts.append(current)

            diff_str = str(diff).lower() if diff else "intermediate"
            cat_str = str(cat) if cat else ""

            current = {
                "id": slugify(str(name)),
                "name": str(name),
                "cat": cat_str,
                "subcat": str(subcat) if subcat else "",
                "dur": int(dur) if dur else 0,
                "diff": diff_str,
                "emoji": WORKOUT_EMOJI.get(cat_str, "💪"),
                "warmup": [],
                "main": [],
                "cooldown": [],
            }
        elif phase and exercise and current is not None:
            # Exercise row
            entry = {
                "exercise": str(exercise),
                "sets": str(sets) if sets else "1",
                "reps": str(reps) if reps else "",
            }
            phase_lower = str(phase).lower().replace("-", "").replace(" ", "")
            if phase_lower == "warmup":
                current["warmup"].append(entry)
            elif phase_lower == "main":
                current["main"].append(entry)
            elif phase_lower == "cooldown":
                current["cooldown"].append(entry)

    # Don't forget the last workout
    if current is not None:
        workouts.append(current)

    return workouts


def build_workouts():
    """Parse all workout detail sheets from the Excel."""
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

    all_workouts = []
    for sheet_name in WORKOUT_SHEETS:
        ws = wb[sheet_name]
        sheet_workouts = parse_workout_sheet(ws)
        all_workouts.extend(sheet_workouts)
        print(f"  {sheet_name}: {len(sheet_workouts)} workouts")

    # Deduplicate by ID (in case of overlap)
    seen = set()
    unique = []
    for wk in all_workouts:
        if wk["id"] not in seen:
            seen.add(wk["id"])
            unique.append(wk)
        else:
            print(f"  WARNING: duplicate workout ID '{wk['id']}' ({wk['name']}) — skipping")

    return unique


def main():
    print("Building exercises from app data...")
    exercises = build_exercises()
    print(f"  {len(exercises)} exercises")

    print("\nParsing workouts from Excel...")
    workouts = build_workouts()
    print(f"  {len(workouts)} workouts total (excluding Challenge Programs)")

    # Write libo-data.js
    print(f"\nWriting {OUTPUT_LIBO_DATA}...")
    exercises_json = json.dumps(exercises, ensure_ascii=False, separators=(",", ": "))
    workouts_json = json.dumps(workouts, ensure_ascii=False, separators=(",", ": "))

    with open(OUTPUT_LIBO_DATA, "w") as f:
        f.write("// Libo – Full Exercise & Workout Library\n")
        f.write("// Auto-generated from Exercise_Library_v5_Workouts.xlsx\n")
        f.write(f"// {len(exercises)} exercises · {len(workouts)} workouts\n")
        f.write("\n")
        f.write(f"const LIBO_EXERCISES = {exercises_json};\n")
        f.write("\n")
        f.write(f"const LIBO_WORKOUTS = {workouts_json};\n")

    # Write react-app/public/exercises.json
    print(f"Writing {OUTPUT_EXERCISES_JSON}...")
    with open(OUTPUT_EXERCISES_JSON, "w") as f:
        json.dump(exercises, f, ensure_ascii=False, separators=(",", ": "))

    # Write react-app/public/workouts.json
    print(f"Writing {OUTPUT_WORKOUTS_JSON}...")
    with open(OUTPUT_WORKOUTS_JSON, "w") as f:
        json.dump(workouts, f, ensure_ascii=False, separators=(",", ": "))

    print("\nDone!")
    print(f"  Exercises: {len(exercises)}")
    print(f"  Workouts: {len(workouts)}")


if __name__ == "__main__":
    main()
